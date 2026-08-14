from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook

from commodity_forecasting.phase1 import readiness
from commodity_forecasting.phase1.contracts import (
    EXPECTED_PERIOD_COUNT,
    EXPECTED_PERIOD_END,
    EXPECTED_PERIOD_START,
    EXPECTED_RAW_SHA256,
    PUBLICATION_POLICY,
    READER_REQUIREMENT,
    REQUIRED_PASS_CHECKS,
    TARGET_COLUMN,
    WORKSHEET_NAME,
    reader_version_supported,
)
from commodity_forecasting.phase1.evidence import (
    EvidenceError,
    assert_no_secret_material,
    assert_finding_matches_record,
    sanitize_diagnostic,
    validate_evidence_record,
    write_evidence,
    write_finding,
)
from commodity_forecasting.phase1.paths import (
    dependency_readiness_evidence_path,
    dependency_readiness_finding_path,
    model_ready_root,
    phase1_evidence_dir,
    phase1_findings_dir,
    raw_workbook_path,
    source_metadata_path,
    standardized_root,
)
from commodity_forecasting.phase1.readiness import (
    CommandResult,
    DuplicateTargetColumnError,
    MonthlyTimestampError,
    ReaderUnavailableError,
    ReaderVersionError,
    RoadmapConsistencyError,
    TargetColumnMissingError,
    WorkbookHashMismatchError,
    WorkbookMissingError,
    WorkbookOpenError,
    WorksheetMissingError,
    assert_latest_run_roadmap_consistency,
    child_environment,
    installed_dependency_declared,
    parse_month_token,
    probe_workbook,
    reconcile_roadmap,
    verify_clean_install,
)


REPO_ROOT = Path(__file__).resolve().parents[2]


def _valid_record(*, classification: str = "pass") -> dict[str, object]:
    checks = {name: True for name in REQUIRED_PASS_CHECKS}
    stages = [
        {"stage": stage, "status": "pass", "exit_code": 0, "sanitized_command": stage}
        for stage in (
            "create_environment",
            "install_project",
            "pip_check",
            "probe_installed",
            "cleanup",
        )
    ]
    record: dict[str, object] = {
        "schema_version": 1,
        "task_id": "P1-01",
        "run_id": "P1-01-20260814T000000Z",
        "timestamp_utc": "2026-08-14T00:00:00Z",
        "classification": classification,
        "mode": "clean-install-readiness",
        "controller_mode": "verify-clean-install",
        "child_mode": "probe-installed",
        "sanitized_command": "python -m commodity_forecasting.phase1.readiness verify-clean-install --repo-root /repo",
        "repo_root": "/repo",
        "stage_outcomes": stages,
        "failed_stage": None,
        "child_exit_code": 0,
        "reader_requirement": READER_REQUIREMENT,
        "resolved_openpyxl_version": "3.1.5",
        "installed_project_version": "0.1.0",
        "python_version": "3.13.0",
        "working_directory_class": "outside_repository",
        "install_mode": "non_editable",
        "pythonpath_state": "absent",
        "host_pythonpath_state": "explicit_repo_src",
        "expected_raw_sha256": EXPECTED_RAW_SHA256,
        "raw_sha256_before": EXPECTED_RAW_SHA256,
        "raw_sha256_after": EXPECTED_RAW_SHA256,
        "sheet_name": WORKSHEET_NAME,
        "target_column": TARGET_COLUMN,
        "header_row": 5,
        "target_column_index": 13,
        "period_start": EXPECTED_PERIOD_START,
        "period_end": EXPECTED_PERIOD_END,
        "period_count": EXPECTED_PERIOD_COUNT,
        "publication_policy": PUBLICATION_POLICY.as_dict(),
        "checks": checks,
        "errors": [],
        "cleanup_status": "pass",
        "artifact_paths": [
            "docs/findings/phase1/evidence/dependency_readiness.json",
            "docs/findings/phase1/dependency_readiness.md",
        ],
        "phase0_evidence": [
            "docs/findings/phase0/evidence/roadmap_exit.json",
            "docs/findings/phase0/evidence/world_bank_pink_sheet_availability.json",
        ],
    }
    if classification != "pass":
        checks["install_non_editable"] = False
        stages[1] = {
            "stage": "install_project",
            "status": "blocked",
            "exit_code": 1,
            "sanitized_command": "pip install /repo",
        }
        record["failed_stage"] = "install_project"
        record["child_mode"] = None
        record["child_exit_code"] = None
        record["resolved_openpyxl_version"] = "unknown"
        record["installed_project_version"] = "unknown"
        record["errors"] = ["project installation failed"]
        record["cleanup_status"] = "pass"
    return record


def _make_workbook(
    path: Path,
    *,
    sheet_name: str = WORKSHEET_NAME,
    target_headers: int = 1,
    periods: tuple[str, ...] = ("1960M01", "1960M02"),
) -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for column in range(1, target_headers + 1):
        sheet.cell(row=4, column=column + 1, value=TARGET_COLUMN)
    sheet.cell(row=5, column=2, value="($/kg)")
    for row, period in enumerate(periods, start=6):
        sheet.cell(row=row, column=1, value=period)
        sheet.cell(row=row, column=2, value=1.0)
    workbook.save(path)
    workbook.close()
    return readiness.sha256_file(path)


def test_dependency_declaration_contains_only_authorized_reader() -> None:
    text = (REPO_ROOT / "pyproject.toml").read_text(encoding="utf-8")
    match = re.search(r"(?m)^dependencies\s*=\s*\[(.*?)\]$", text)

    assert match is not None
    assert match.group(1).strip() == f'"{READER_REQUIREMENT}"'


def test_installed_dependency_declaration_is_verified(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        readiness.importlib.metadata,
        "requires",
        lambda _: ["openpyxl<4,>=3.1"],
    )
    assert installed_dependency_declared() is True

    monkeypatch.setattr(readiness.importlib.metadata, "requires", lambda _: ["openpyxl>=3.0"])
    assert installed_dependency_declared() is False

    monkeypatch.setattr(readiness.importlib.metadata, "requires", lambda _: ["openpyxl>=3.10,<40"])
    assert installed_dependency_declared() is False


@pytest.mark.parametrize("version", ["3.1", "3.1.0", "3.99.4"])
def test_reader_version_contract_accepts_supported_final_releases(version: str) -> None:
    assert reader_version_supported(version) is True


@pytest.mark.parametrize("version", ["3.0.99", "4.0", "3.1rc1", "3.1.0.post1", "unknown"])
def test_reader_version_contract_rejects_out_of_range_or_nonfinal_releases(version: str) -> None:
    assert reader_version_supported(version) is False


def test_phase1_paths_resolve_without_creating_directories(tmp_path: Path) -> None:
    assert raw_workbook_path() == (
        REPO_ROOT / "data" / "raw" / "world_bank" / "pink_sheet" / "CMO-Historical-Data-Monthly.xlsx"
    )
    assert dependency_readiness_evidence_path() == (
        REPO_ROOT / "docs" / "findings" / "phase1" / "evidence" / "dependency_readiness.json"
    )
    assert dependency_readiness_finding_path() == (
        REPO_ROOT / "docs" / "findings" / "phase1" / "dependency_readiness.md"
    )
    assert raw_workbook_path(tmp_path) == (
        tmp_path / "data" / "raw" / "world_bank" / "pink_sheet" / "CMO-Historical-Data-Monthly.xlsx"
    )
    assert source_metadata_path(tmp_path).name == "source_metadata.json"
    assert phase1_findings_dir(tmp_path) == tmp_path / "docs" / "findings" / "phase1"
    assert phase1_evidence_dir(tmp_path) == tmp_path / "docs" / "findings" / "phase1" / "evidence"
    assert standardized_root(tmp_path) == tmp_path / "data" / "standardized"
    assert model_ready_root(tmp_path) == tmp_path / "data" / "model_ready"
    assert not (tmp_path / "docs").exists()
    assert not (tmp_path / "data").exists()


@pytest.mark.parametrize(
    ("token", "expected"),
    [
        ("1960M01", date(1960, 1, 1)),
        ("2000M12", date(2000, 12, 1)),
        ("2026M07", date(2026, 7, 1)),
    ],
)
def test_monthly_timestamp_parsing(token: str, expected: date) -> None:
    assert parse_month_token(token) == expected


@pytest.mark.parametrize(
    "token",
    ["1960M00", "1960M13", "1960M1", "1960-01", "M01", " 1960M01", "1960M01 ", "", None, 196001],
)
def test_monthly_timestamp_parsing_rejects_invalid_values(token: object) -> None:
    with pytest.raises(MonthlyTimestampError):
        parse_month_token(token)


def test_readiness_cli_requires_an_explicit_absolute_repo_root() -> None:
    with pytest.raises(SystemExit):
        readiness._parser().parse_args(["verify-clean-install"])
    assert readiness.main(["verify-clean-install", "--repo-root", "relative/repo"]) == 1


def test_reader_availability_and_version_errors_are_explicit(monkeypatch: pytest.MonkeyPatch) -> None:
    def missing_reader(_: str) -> object:
        raise ModuleNotFoundError("openpyxl")

    monkeypatch.setattr(readiness.importlib, "import_module", missing_reader)
    with pytest.raises(ReaderUnavailableError):
        readiness._load_reader()

    monkeypatch.setattr(readiness.importlib.metadata, "version", lambda _: "4.0.0")
    with pytest.raises(ReaderVersionError):
        readiness._reader_version()


def test_publication_policy_is_conservative_and_not_vintage_real_time() -> None:
    assert PUBLICATION_POLICY.evaluation_label == "revised_workbook_pseudo_real_time"
    assert PUBLICATION_POLICY.historical_release_timestamps_available is False
    assert PUBLICATION_POLICY.historical_vintages_available is False
    assert PUBLICATION_POLICY.is_available(date(2026, 6, 1), date(2026, 7, 1)) is True
    assert PUBLICATION_POLICY.is_available(date(2026, 7, 1), date(2026, 7, 1)) is False
    assert "preserved workbook" in PUBLICATION_POLICY.limitation
    assert "vintage-real-time" in PUBLICATION_POLICY.prohibited_claim


@pytest.mark.parametrize("classification", ["pass", "fail", "blocked", "unsupported"])
def test_evidence_schema_classifications(classification: str) -> None:
    validate_evidence_record(_valid_record(classification=classification))


def test_evidence_schema_rejects_incomplete_pass_and_secrets() -> None:
    incomplete = _valid_record()
    incomplete["checks"]["raw_hash_after_matches"] = False  # type: ignore[index]
    with pytest.raises(EvidenceError):
        validate_evidence_record(incomplete)

    secret = _valid_record(classification="blocked")
    secret["errors"] = ["api_key=abc123"]
    with pytest.raises(EvidenceError):
        validate_evidence_record(secret)

    for unsupported_version in ("2.9.9", "4.0.0", "3.1rc1"):
        unsupported = _valid_record()
        unsupported["resolved_openpyxl_version"] = unsupported_version
        with pytest.raises(EvidenceError):
            validate_evidence_record(unsupported)

    wrong_requirement = _valid_record()
    wrong_requirement["reader_requirement"] = "openpyxl>=3"
    with pytest.raises(EvidenceError):
        validate_evidence_record(wrong_requirement)


@pytest.mark.parametrize(
    ("mutation", "value"),
    [
        ("classification", "maybe"),
        ("timestamp_utc", "not-a-timestamp"),
        ("artifact_paths", "not-a-list"),
        ("publication_policy", None),
    ],
)
def test_evidence_schema_rejects_invalid_fields(mutation: str, value: object) -> None:
    record = _valid_record(classification="blocked")
    record[mutation] = value
    with pytest.raises(EvidenceError):
        validate_evidence_record(record)

    missing = _valid_record(classification="blocked")
    del missing[mutation]
    with pytest.raises(EvidenceError):
        validate_evidence_record(missing)


def test_evidence_secret_safe_and_atomic_finding_mirror(tmp_path: Path) -> None:
    record = _valid_record()
    evidence_path = tmp_path / "evidence" / "dependency_readiness.json"
    finding_path = tmp_path / "dependency_readiness.md"

    write_evidence(evidence_path, record)
    write_finding(finding_path, record)

    persisted = json.loads(evidence_path.read_text(encoding="utf-8"))
    validate_evidence_record(persisted)
    assert_finding_matches_record(finding_path, persisted)
    finding = finding_path.read_text(encoding="utf-8")
    assert "## Stage outcomes" in finding
    assert "Acceptance command:" in finding
    assert "Controller / child mode:" in finding
    assert "Cleanup:" in finding
    assert not list(tmp_path.rglob("*.tmp"))

    finding_path.write_text(finding.replace("1960M01", "1960M02"), encoding="utf-8")
    with pytest.raises(EvidenceError, match="finding does not match"):
        assert_finding_matches_record(finding_path, persisted)


@pytest.mark.parametrize(
    ("diagnostic", "secret"),
    [
        ('{"token":"abc+/="}', "abc+/="),
        ("https://user:p%40ss@example.com/simple", "user:p%40ss"),
        ("https://example.com/simple?api_key=abc+/=", "abc+/="),
        ("pip --password s3cr3t install", "s3cr3t"),
        ("Authorization: Bearer abc+/=", "abc+/="),
        ("api_key='abc+/='", "abc+/="),
    ],
)
def test_diagnostic_redaction_covers_common_credential_forms(
    diagnostic: str,
    secret: str,
) -> None:
    with pytest.raises(EvidenceError, match="secret-like material"):
        assert_no_secret_material(diagnostic)

    sanitized = sanitize_diagnostic(diagnostic)
    assert secret not in sanitized
    assert "[REDACTED]" in sanitized
    assert_no_secret_material(sanitized)


def test_probe_workbook_detects_contract_and_preserves_hash(tmp_path: Path) -> None:
    workbook_path = tmp_path / "valid.xlsx"
    expected_hash = _make_workbook(workbook_path)

    observation = probe_workbook(workbook_path, expected_sha256=expected_hash)

    assert observation.sheet_name == WORKSHEET_NAME
    assert observation.target_column == TARGET_COLUMN
    assert observation.header_row == 4
    assert observation.target_column_index == 2
    assert observation.period_start == "1960M01"
    assert observation.period_end == "1960M02"
    assert observation.period_count == 2
    assert observation.sha256_before == observation.sha256_after == expected_hash


def test_real_workbook_contract_and_raw_hash_are_unchanged() -> None:
    path = raw_workbook_path(REPO_ROOT)
    before = readiness.sha256_file(path)

    observation = probe_workbook(path, expected_sha256=EXPECTED_RAW_SHA256)

    assert observation.period_start == EXPECTED_PERIOD_START
    assert observation.period_end == EXPECTED_PERIOD_END
    assert observation.period_count == EXPECTED_PERIOD_COUNT
    assert readiness.sha256_file(path) == before == EXPECTED_RAW_SHA256


def test_probe_workbook_explicit_contract_errors(tmp_path: Path) -> None:
    missing = tmp_path / "missing.xlsx"
    with pytest.raises(WorkbookMissingError):
        probe_workbook(missing, expected_sha256="0" * 64)

    wrong_sheet = tmp_path / "wrong-sheet.xlsx"
    wrong_sheet_hash = _make_workbook(wrong_sheet, sheet_name="Other")
    with pytest.raises(WorksheetMissingError):
        probe_workbook(wrong_sheet, expected_sha256=wrong_sheet_hash)

    no_target = tmp_path / "no-target.xlsx"
    no_target_hash = _make_workbook(no_target, target_headers=0)
    with pytest.raises(TargetColumnMissingError):
        probe_workbook(no_target, expected_sha256=no_target_hash)

    duplicate = tmp_path / "duplicate.xlsx"
    duplicate_hash = _make_workbook(duplicate, target_headers=2)
    with pytest.raises(DuplicateTargetColumnError):
        probe_workbook(duplicate, expected_sha256=duplicate_hash)

    malformed = tmp_path / "malformed.xlsx"
    malformed_hash = _make_workbook(malformed, periods=("1960M01", "1960M13"))
    with pytest.raises(MonthlyTimestampError):
        probe_workbook(malformed, expected_sha256=malformed_hash)

    with pytest.raises(WorkbookHashMismatchError):
        probe_workbook(wrong_sheet, expected_sha256="0" * 64)

    corrupt = tmp_path / "corrupt.xlsx"
    corrupt.write_bytes(b"not an xlsx archive")
    with pytest.raises(WorkbookOpenError):
        probe_workbook(corrupt, expected_sha256=readiness.sha256_file(corrupt))


def test_probe_closes_workbook_and_post_read_hash_mismatch_takes_precedence(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    workbook_path = tmp_path / "fixture.xlsx"
    workbook_path.write_bytes(b"fixture")
    expected_hash = readiness.sha256_file(workbook_path)

    class FailingWorksheet:
        def iter_rows(self, *, values_only: bool) -> object:
            assert values_only is True
            raise RuntimeError("reader failed after open")

    class FailingWorkbook:
        sheetnames = [WORKSHEET_NAME]
        closed = False

        def __getitem__(self, _: str) -> FailingWorksheet:
            return FailingWorksheet()

        def close(self) -> None:
            self.closed = True

    workbook = FailingWorkbook()

    class FakeReader:
        @staticmethod
        def load_workbook(path: Path, *, read_only: bool, data_only: bool) -> FailingWorkbook:
            assert path == workbook_path
            assert read_only is True and data_only is True
            return workbook

    monkeypatch.setattr(readiness, "_load_reader", lambda: FakeReader)
    hashes = iter((expected_hash, "0" * 64))
    monkeypatch.setattr(readiness, "sha256_file", lambda _: next(hashes))

    with pytest.raises(WorkbookHashMismatchError):
        probe_workbook(workbook_path, expected_sha256=expected_hash)
    assert workbook.closed is True


def test_child_preflight_fails_before_reader_version_check(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    phase0 = tmp_path / "docs" / "findings" / "phase0" / "evidence" / "roadmap_exit.json"
    phase0.parent.mkdir(parents=True)
    phase0.write_text('{"classification": "pass"}\n', encoding="utf-8")
    metadata = tmp_path / "data" / "raw" / "world_bank" / "pink_sheet" / "source_metadata.json"
    metadata.parent.mkdir(parents=True)
    metadata.write_text(json.dumps({"sha256": EXPECTED_RAW_SHA256}), encoding="utf-8")
    monkeypatch.setattr(readiness, "installed_dependency_declared", lambda: True)

    def unexpected_reader_check() -> str:
        raise AssertionError("reader version must not run before deterministic preflight")

    monkeypatch.setattr(readiness, "_reader_version", unexpected_reader_check)
    with pytest.raises(WorkbookMissingError):
        readiness._child_probe(tmp_path)


def test_child_environment_removes_pythonpath() -> None:
    environment = child_environment({"PYTHONPATH": "/repo/src", "PATH": "/bin"})

    assert "PYTHONPATH" not in environment
    assert environment["PATH"] == "/bin"


def _prepare_controller_repo(root: Path, *, checked: bool = True) -> None:
    (root / "src").mkdir(parents=True)
    (root / "pyproject.toml").write_text(
        "[build-system]\nrequires = [\"setuptools>=68\"]\nbuild-backend = \"setuptools.build_meta\"\n",
        encoding="utf-8",
    )
    roadmap = root / "docs" / "roadmap.md"
    roadmap.parent.mkdir(parents=True)
    marker = "x" if checked else " "
    roadmap.write_text(f"- [{marker}] **P1-01 — Dependency/readiness contract.**\n", encoding="utf-8")


def _child_result(classification: str) -> dict[str, object]:
    result: dict[str, object] = {
        "classification": classification,
        "checks": {name: True for name in REQUIRED_PASS_CHECKS},
        "errors": [] if classification == "pass" else ["deterministic workbook contract mismatch"],
        "resolved_openpyxl_version": "3.1.5",
        "installed_project_version": "0.1.0",
        "python_version": "3.13.0",
        "working_directory_class": "outside_repository",
        "install_mode": "non_editable",
        "pythonpath_state": "absent",
        "sha256_before": EXPECTED_RAW_SHA256,
        "sha256_after": EXPECTED_RAW_SHA256,
        "sheet_name": WORKSHEET_NAME,
        "target_column": TARGET_COLUMN,
        "header_row": 5,
        "target_column_index": 13,
        "period_start": EXPECTED_PERIOD_START,
        "period_end": EXPECTED_PERIOD_END,
        "period_count": EXPECTED_PERIOD_COUNT,
    }
    return result


def _runner_with_child(classification: str) -> readiness.CommandRunner:
    def fake_runner(
        stage: str,
        command: list[str],
        *,
        cwd: Path,
        env: dict[str, str],
    ) -> CommandResult:
        del cwd
        assert "PYTHONPATH" not in env
        if stage == "probe_installed":
            result_path = Path(command[command.index("--result-path") + 1])
            result_path.write_text(json.dumps(_child_result(classification)), encoding="utf-8")
            return CommandResult(returncode=0 if classification == "pass" else 1, stdout="", stderr="")
        return CommandResult(returncode=0, stdout="", stderr="")

    return fake_runner


def test_controller_preserves_deterministic_child_fail(tmp_path: Path) -> None:
    _prepare_controller_repo(tmp_path)
    temporary_root = tmp_path / "controller"

    record = verify_clean_install(
        tmp_path,
        command_runner=_runner_with_child("fail"),
        temporary_root_factory=lambda: temporary_root,
    )

    assert record["classification"] == "fail"
    assert record["failed_stage"] == "probe_installed"
    assert record["errors"] == ["deterministic workbook contract mismatch"]
    assert "- [ ] **P1-01" in (tmp_path / "docs" / "roadmap.md").read_text(encoding="utf-8")


def test_invalid_child_pass_is_published_as_fail_instead_of_leaving_stale_pass(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _prepare_controller_repo(tmp_path)
    monkeypatch.setenv("PYTHONPATH", str(tmp_path / "src"))
    temporary_root = tmp_path / "controller"

    def fake_runner(
        stage: str,
        command: list[str],
        *,
        cwd: Path,
        env: dict[str, str],
    ) -> CommandResult:
        del cwd, env
        if stage == "probe_installed":
            invalid_pass = _child_result("pass")
            invalid_pass["resolved_openpyxl_version"] = "4.0.0"
            result_path = Path(command[command.index("--result-path") + 1])
            result_path.write_text(json.dumps(invalid_pass), encoding="utf-8")
        return CommandResult(returncode=0, stdout="", stderr="")

    record = verify_clean_install(
        tmp_path,
        command_runner=fake_runner,
        temporary_root_factory=lambda: temporary_root,
    )

    assert record["classification"] == "fail"
    assert record["failed_stage"] == "evidence_validation"
    assert record["errors"] == ["evidence validation failed: pass requires an openpyxl version satisfying >=3.1,<4"]
    persisted = json.loads(
        (tmp_path / "docs" / "findings" / "phase1" / "evidence" / "dependency_readiness.json").read_text(
            encoding="utf-8"
        )
    )
    validate_evidence_record(persisted)
    assert "- [ ] **P1-01" in (tmp_path / "docs" / "roadmap.md").read_text(encoding="utf-8")


def test_cleanup_failure_retains_controller_diagnostic(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    _prepare_controller_repo(tmp_path)
    temporary_root = tmp_path / "controller"

    def fail_cleanup(_: Path) -> None:
        raise OSError("cleanup fixture failure")

    monkeypatch.setattr(readiness.shutil, "rmtree", fail_cleanup)
    record = verify_clean_install(
        tmp_path,
        command_runner=_runner_with_child("pass"),
        temporary_root_factory=lambda: temporary_root,
    )

    assert record["classification"] == "blocked"
    assert record["failed_stage"] == "cleanup"
    assert record["errors"] == ["cleanup failed: cleanup fixture failure"]
    assert record["cleanup_status"] == "blocked"


def test_secret_bearing_install_failure_replaces_stale_pass_safely(tmp_path: Path) -> None:
    _prepare_controller_repo(tmp_path)
    temporary_root = tmp_path / "controller"

    passing = _valid_record()
    write_evidence(
        tmp_path / "docs" / "findings" / "phase1" / "evidence" / "dependency_readiness.json",
        passing,
    )
    write_finding(tmp_path / "docs" / "findings" / "phase1" / "dependency_readiness.md", passing)

    def fake_runner(
        stage: str,
        command: list[str],
        *,
        cwd: Path,
        env: dict[str, str],
    ) -> CommandResult:
        del command, cwd, env
        if stage == "install_project":
            return CommandResult(returncode=1, stdout="", stderr="download failed api_key=abc123")
        return CommandResult(returncode=0, stdout="", stderr="")

    record = verify_clean_install(
        tmp_path,
        command_runner=fake_runner,
        temporary_root_factory=lambda: temporary_root,
    )

    persisted_text = (
        tmp_path / "docs" / "findings" / "phase1" / "evidence" / "dependency_readiness.json"
    ).read_text(encoding="utf-8")
    assert record["classification"] == "blocked"
    assert "abc123" not in persisted_text
    assert "[REDACTED]" in persisted_text
    assert "- [ ] **P1-01" in (tmp_path / "docs" / "roadmap.md").read_text(encoding="utf-8")


@pytest.mark.parametrize(
    ("failed_stage", "expected_calls"),
    [
        ("create_environment", ["create_environment"]),
        ("pip_check", ["create_environment", "install_project", "pip_check"]),
        ("probe_installed", ["create_environment", "install_project", "pip_check", "probe_installed"]),
    ],
)
def test_controller_stage_failures_stop_and_publish_blocked(
    tmp_path: Path, failed_stage: str, expected_calls: list[str]
) -> None:
    _prepare_controller_repo(tmp_path)
    calls: list[str] = []
    temporary_root = tmp_path / "controller"

    def fake_runner(
        stage: str,
        command: list[str],
        *,
        cwd: Path,
        env: dict[str, str],
    ) -> CommandResult:
        del command, cwd, env
        calls.append(stage)
        if stage == failed_stage:
            return CommandResult(returncode=1, stdout="", stderr=f"{stage} fixture failure")
        return CommandResult(returncode=0, stdout="", stderr="")

    record = verify_clean_install(
        tmp_path,
        command_runner=fake_runner,
        temporary_root_factory=lambda: temporary_root,
    )

    assert record["classification"] == "blocked"
    assert record["failed_stage"] == failed_stage
    assert calls == expected_calls
    assert record["cleanup_status"] == "pass"
    assert "- [ ] **P1-01" in (tmp_path / "docs" / "roadmap.md").read_text(encoding="utf-8")


def test_install_failure_publishes_blocked(tmp_path: Path) -> None:
    calls: list[str] = []

    def fake_runner(
        stage: str,
        command: list[str],
        *,
        cwd: Path,
        env: dict[str, str],
    ) -> CommandResult:
        del command, cwd, env
        calls.append(stage)
        if stage == "install_project":
            return CommandResult(returncode=1, stdout="", stderr="package index unavailable")
        return CommandResult(returncode=0, stdout="", stderr="")

    controller_root = tmp_path / "controller"
    controller_root.mkdir()
    (tmp_path / "pyproject.toml").write_text(
        "[build-system]\nrequires = [\"setuptools>=68\"]\nbuild-backend = \"setuptools.build_meta\"\n",
        encoding="utf-8",
    )
    (tmp_path / "src").mkdir()
    roadmap = tmp_path / "docs" / "roadmap.md"
    roadmap.parent.mkdir(parents=True)
    roadmap.write_text("- [x] **P1-01 — Dependency/readiness contract.**\n", encoding="utf-8")

    record = verify_clean_install(
        tmp_path,
        command_runner=fake_runner,
        temporary_root_factory=lambda: controller_root,
    )

    assert record["classification"] == "blocked"
    assert record["failed_stage"] == "install_project"
    assert calls == ["create_environment", "install_project"]
    assert record["cleanup_status"] == "pass"
    assert not controller_root.exists()
    assert "- [ ] **P1-01" in roadmap.read_text(encoding="utf-8")
    persisted = json.loads(
        (tmp_path / "docs" / "findings" / "phase1" / "evidence" / "dependency_readiness.json").read_text(
            encoding="utf-8"
        )
    )
    validate_evidence_record(persisted)


def test_host_bootstrap_starts_from_uninstalled_src_checkout(tmp_path: Path) -> None:
    repository = tmp_path / "repository"
    package_root = repository / "src" / "commodity_forecasting"
    package_root.parent.mkdir(parents=True)
    shutil.copytree(REPO_ROOT / "src" / "commodity_forecasting", package_root)
    roadmap = repository / "docs" / "roadmap.md"
    roadmap.parent.mkdir(parents=True)
    roadmap.write_text("- [x] **P1-01 — Dependency/readiness contract.**\n", encoding="utf-8")
    outside = tmp_path / "outside"
    outside.mkdir()
    environment = dict(os.environ)
    environment["PYTHONPATH"] = str(repository / "src")

    result = subprocess.run(
        [
            sys.executable,
            "-m",
            "commodity_forecasting.phase1.readiness",
            "verify-clean-install",
            "--repo-root",
            str(repository),
        ],
        cwd=outside,
        env=environment,
        check=False,
        capture_output=True,
        text=True,
    )

    assert result.returncode == 1
    evidence = json.loads(
        (repository / "docs" / "findings" / "phase1" / "evidence" / "dependency_readiness.json").read_text(
            encoding="utf-8"
        )
    )
    assert evidence["classification"] == "blocked"
    assert evidence["checks"]["host_bootstrap_explicit"] is True
    assert evidence["host_pythonpath_state"] == "explicit_repo_src"
    assert evidence["sanitized_command"].startswith(f"PYTHONPATH={repository / 'src'} {sys.executable}")
    assert [stage["stage"] for stage in evidence["stage_outcomes"]] == [
        "create_environment",
        "install_project",
        "cleanup",
    ]
    assert "- [ ] **P1-01" in roadmap.read_text(encoding="utf-8")


def test_latest_run_roadmap_consistency(tmp_path: Path) -> None:
    roadmap = tmp_path / "docs" / "roadmap.md"
    roadmap.parent.mkdir(parents=True)
    roadmap.write_text("- [ ] **P1-01 — Dependency/readiness contract.**\n", encoding="utf-8")
    evidence_path = tmp_path / "docs" / "findings" / "phase1" / "evidence" / "dependency_readiness.json"
    finding_path = tmp_path / "docs" / "findings" / "phase1" / "dependency_readiness.md"

    passing = _valid_record()
    write_evidence(evidence_path, passing)
    write_finding(finding_path, passing)
    reconcile_roadmap(tmp_path, passing)
    assert_latest_run_roadmap_consistency(tmp_path)
    assert "- [x] **P1-01" in roadmap.read_text(encoding="utf-8")

    blocked = _valid_record(classification="blocked")
    write_evidence(evidence_path, blocked)
    write_finding(finding_path, blocked)
    reconcile_roadmap(tmp_path, blocked)
    assert_latest_run_roadmap_consistency(tmp_path)
    assert "- [ ] **P1-01" in roadmap.read_text(encoding="utf-8")

    roadmap.write_text(
        roadmap.read_text(encoding="utf-8")
        + "- [x] **P1-01 — Conflicting duplicate.**\n",
        encoding="utf-8",
    )
    with pytest.raises(RoadmapConsistencyError, match="does not match"):
        assert_latest_run_roadmap_consistency(tmp_path)
