from __future__ import annotations

import csv
import json
import os
import shutil
import subprocess
import sys
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from commodity_forecasting.phase1 import target_pipeline
from commodity_forecasting.phase1.contracts import (
    EXPECTED_PERIOD_COUNT,
    EXPECTED_PERIOD_END,
    EXPECTED_PERIOD_START,
    EXPECTED_RAW_SHA256,
    PUBLICATION_POLICY,
    TARGET_COLUMN,
    WORKSHEET_NAME,
)
from commodity_forecasting.phase1.readiness import sha256_file
from commodity_forecasting.phase1.target_pipeline import (
    CSV_FIELDS,
    EVIDENCE_RELATIVE_PATH,
    FINDING_RELATIVE_PATH,
    MODEL_READY_RELATIVE_PATH,
    STANDARDIZED_RELATIVE_PATH,
    UNIQUE_ID,
    DependencyGateError,
    InvalidTargetValueError,
    InvalidTimestampError,
    PublicationError,
    RoadmapEligibilityError,
    TargetEvidenceError,
    TargetRow,
    TargetSelectionError,
    assert_roadmap_completion_eligible,
    assert_roadmap_consistent,
    extract_target,
    publish_target_pipeline,
    serialize_csv,
    serialize_numeric,
    validate_published_state,
    validate_rows,
    validate_target_evidence,
)

REPO_ROOT = Path(__file__).resolve().parents[2]
RAW_RELATIVE_PATH = Path(
    "data/raw/world_bank/pink_sheet/CMO-Historical-Data-Monthly.xlsx"
)
FIXED_NOW = datetime(2026, 8, 14, 8, 30, tzinfo=timezone.utc)


@pytest.fixture(scope="module")
def real_observation() -> target_pipeline.TargetObservation:
    return extract_target(REPO_ROOT / RAW_RELATIVE_PATH)


def _copy(path: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(path, destination)


def _isolated_root(tmp_path: Path, name: str = "repo") -> Path:
    root = tmp_path / name
    _copy(REPO_ROOT / RAW_RELATIVE_PATH, root / RAW_RELATIVE_PATH)
    _copy(
        REPO_ROOT / "docs/findings/phase1/evidence/dependency_readiness.json",
        root / "docs/findings/phase1/evidence/dependency_readiness.json",
    )
    _copy(
        REPO_ROOT / "docs/findings/phase1/dependency_readiness.md",
        root / "docs/findings/phase1/dependency_readiness.md",
    )
    _copy(REPO_ROOT / "docs/roadmap.md", root / "docs/roadmap.md")
    return root


def _workbook(path: Path, *, header: str = TARGET_COLUMN, period: object = "1960M13") -> str:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = WORKSHEET_NAME
    sheet.cell(row=1, column=2, value=header)
    sheet.cell(row=2, column=1, value=period)
    sheet.cell(row=2, column=2, value=1.14)
    workbook.save(path)
    workbook.close()
    return sha256_file(path)


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _valid_record(
    observation: target_pipeline.TargetObservation,
) -> dict[str, object]:
    csv_bytes = serialize_csv(observation.rows)
    digest = target_pipeline.sha256_bytes(csv_bytes)
    return target_pipeline._build_record(
        observation,
        standardized_sha256=digest,
        model_ready_sha256=digest,
        now=FIXED_NOW,
    )


def test_exact_target_selection_and_real_monthly_extent(
    real_observation: target_pipeline.TargetObservation,
) -> None:
    assert real_observation.header_row == 5
    assert real_observation.target_column_index == 13
    assert real_observation.period_start == EXPECTED_PERIOD_START
    assert real_observation.period_end == EXPECTED_PERIOD_END
    assert len(real_observation.rows) == EXPECTED_PERIOD_COUNT == 799
    assert real_observation.rows[0].ds == date(1960, 1, 1)
    assert real_observation.rows[-1].ds == date(2026, 7, 1)
    assert real_observation.source_sha256_before == EXPECTED_RAW_SHA256
    assert real_observation.source_sha256_after == EXPECTED_RAW_SHA256


def test_exact_target_selection_preserves_every_observed_price_level(
    real_observation: target_pipeline.TargetObservation,
) -> None:
    workbook = load_workbook(REPO_ROOT / RAW_RELATIVE_PATH, read_only=True, data_only=True)
    try:
        sheet = workbook[WORKSHEET_NAME]
        source_values = [
            row[12]
            for row in sheet.iter_rows(
                min_row=7,
                max_row=6 + EXPECTED_PERIOD_COUNT,
                max_col=13,
                values_only=True,
            )
        ]
    finally:
        workbook.close()
    assert [row.y for row in real_observation.rows] == [
        serialize_numeric(value) for value in source_values
    ]


def test_missing_target_failure_is_explicit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "missing-target.xlsx"
    source_hash = _workbook(path, header="Coffee, Robusta")
    monkeypatch.setattr(target_pipeline, "EXPECTED_RAW_SHA256", source_hash)
    with pytest.raises(TargetSelectionError, match="target column not found"):
        extract_target(path)
    assert sha256_file(path) == source_hash


def test_duplicate_target_failure_is_explicit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    path = tmp_path / "duplicate-target.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = WORKSHEET_NAME
    sheet.cell(row=1, column=2, value=TARGET_COLUMN)
    sheet.cell(row=1, column=3, value=TARGET_COLUMN)
    workbook.save(path)
    workbook.close()
    source_hash = sha256_file(path)
    monkeypatch.setattr(target_pipeline, "EXPECTED_RAW_SHA256", source_hash)
    with pytest.raises(TargetSelectionError, match="exactly once"):
        extract_target(path)


@pytest.mark.parametrize("period", ["1960M00", "1960M13", "1960M01 ", 196001])
def test_invalid_timestamp_failure_is_explicit(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    period: object,
) -> None:
    path = tmp_path / "invalid-timestamp.xlsx"
    source_hash = _workbook(path, period="1960M01")
    workbook = load_workbook(path)
    workbook[WORKSHEET_NAME].cell(row=3, column=1, value=period)
    workbook[WORKSHEET_NAME].cell(row=3, column=2, value=1.15)
    workbook.save(path)
    workbook.close()
    source_hash = sha256_file(path)
    monkeypatch.setattr(target_pipeline, "EXPECTED_RAW_SHA256", source_hash)
    with pytest.raises(InvalidTimestampError, match="invalid monthly timestamp"):
        extract_target(path)
    assert sha256_file(path) == source_hash


@pytest.mark.parametrize(
    ("value", "expected"),
    [(1.14, "1.14"), (1, "1"), (0.0, "0.0"), (-2.5, "-2.5")],
)
def test_numeric_serializer_is_deterministic(value: object, expected: str) -> None:
    assert serialize_numeric(value) == expected
    assert "e" not in expected.lower()
    assert Decimal(expected).is_finite()


@pytest.mark.parametrize("value", [None, True, False, "1.14", float("nan"), float("inf"), -float("inf")])
def test_numeric_serializer_rejects_non_numeric_or_non_finite(value: object) -> None:
    with pytest.raises(InvalidTargetValueError):
        serialize_numeric(value)


def test_monthly_rows_must_be_unique_ordered_and_contiguous(
    real_observation: target_pipeline.TargetObservation,
) -> None:
    validate_rows(real_observation.rows)
    rows = list(real_observation.rows)
    rows[1] = TargetRow(UNIQUE_ID, rows[0].ds, rows[1].y)
    with pytest.raises(InvalidTimestampError, match="unique and strictly ordered"):
        validate_rows(rows)

    rows = list(real_observation.rows)
    rows[1] = TargetRow(UNIQUE_ID, date(1960, 3, 1), rows[1].y)
    with pytest.raises(InvalidTimestampError):
        validate_rows(rows)


def test_target_only_csv_schema_and_numeric_non_null_y(
    real_observation: target_pipeline.TargetObservation,
    tmp_path: Path,
) -> None:
    content = serialize_csv(real_observation.rows)
    assert content.startswith(b"unique_id,ds,y\n")
    assert content.endswith(b"\n")
    assert b"\r" not in content
    assert not content.startswith(b"\xef\xbb\xbf")
    path = tmp_path / "target.csv"
    path.write_bytes(content)
    rows = _read_csv(path)
    assert list(rows[0]) == list(CSV_FIELDS)
    assert all(row["unique_id"] == UNIQUE_ID for row in rows)
    assert all(row["y"] and Decimal(row["y"]).is_finite() for row in rows)
    assert all(set(row) == set(CSV_FIELDS) for row in rows)


def test_compact_evidence_schema_and_publication_policy(
    real_observation: target_pipeline.TargetObservation,
) -> None:
    record = _valid_record(real_observation)
    validate_target_evidence(record)
    assert record["schemas"] == {
        "standardized": ["unique_id", "ds", "y"],
        "model_ready": ["unique_id", "ds", "y"],
    }
    assert record["publication_policy"] == PUBLICATION_POLICY.as_dict()
    checks = record["checks"]
    assert isinstance(checks, dict)
    assert set(checks) == set(target_pipeline.REQUIRED_CHECKS)
    assert all(type(value) is bool for value in checks.values())

    record["recovery_state"] = {}
    with pytest.raises(TargetEvidenceError, match="unsupported evidence fields"):
        validate_target_evidence(record)


@pytest.mark.parametrize(
    ("field", "value", "message"),
    [
        ("source_workbook_path", "not-the-source.xlsx", "canonical source workbook"),
        ("header_row", -1, "canonical header and target index"),
        ("target_column_index", -1, "canonical header and target index"),
    ],
)
def test_compact_evidence_rejects_false_provenance(
    real_observation: target_pipeline.TargetObservation,
    field: str,
    value: object,
    message: str,
) -> None:
    record = _valid_record(real_observation)
    record[field] = value
    with pytest.raises(TargetEvidenceError, match=message):
        validate_target_evidence(record)


def test_compact_evidence_rejects_false_checks_and_altered_schemas(
    real_observation: target_pipeline.TargetObservation,
) -> None:
    false_check = _valid_record(real_observation)
    checks = false_check["checks"]
    assert isinstance(checks, dict)
    checks["evidence_json_written_last"] = False
    with pytest.raises(TargetEvidenceError, match="true checks"):
        validate_target_evidence(false_check)

    wrong_schema = _valid_record(real_observation)
    wrong_schema["schemas"] = {
        "standardized": ["unique_id", "ds", "y", "feature"],
        "model_ready": ["unique_id", "ds", "y"],
    }
    with pytest.raises(TargetEvidenceError, match="target-only"):
        validate_target_evidence(wrong_schema)


def test_current_p101_evidence_must_be_valid_and_pass(tmp_path: Path) -> None:
    root = _isolated_root(tmp_path)
    record = publish_target_pipeline(root, now=FIXED_NOW)
    assert record["checks"]["p101_evidence_valid"] is True
    assert record["checks"]["p101_evidence_pass"] is True
    assert record["checks"]["p101_finding_matches"] is True
    assert record["checks"]["p101_publication_policy_matches"] is True

    gated_root = _isolated_root(tmp_path, "blocked")
    evidence_path = gated_root / "docs/findings/phase1/evidence/dependency_readiness.json"
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    evidence["classification"] = "blocked"
    evidence_path.write_text(json.dumps(evidence), encoding="utf-8")
    with pytest.raises(DependencyGateError, match="not a current canonical pass"):
        publish_target_pipeline(gated_root, now=FIXED_NOW)


def test_determinism_across_two_public_command_runs(tmp_path: Path) -> None:
    roots = [_isolated_root(tmp_path, "run-a"), _isolated_root(tmp_path, "run-b")]
    outputs: list[dict[str, object]] = []
    environment = dict(os.environ)
    environment["PYTHONPATH"] = str(REPO_ROOT / "src")
    for root in roots:
        result = subprocess.run(
            [
                sys.executable,
                "-m",
                "commodity_forecasting.phase1.target_pipeline",
                "publish",
                "--repo-root",
                str(root.resolve()),
            ],
            cwd=REPO_ROOT,
            env=environment,
            check=False,
            capture_output=True,
            text=True,
        )
        assert result.returncode == 0, result.stderr
        outputs.append(json.loads(result.stdout))

    assert outputs[0]["standardized_sha256"] == outputs[1]["standardized_sha256"]
    assert outputs[0]["model_ready_sha256"] == outputs[1]["model_ready_sha256"]
    assert _read_csv(roots[0] / STANDARDIZED_RELATIVE_PATH) == _read_csv(
        roots[1] / STANDARDIZED_RELATIVE_PATH
    )
    assert _read_csv(roots[0] / MODEL_READY_RELATIVE_PATH) == _read_csv(
        roots[1] / MODEL_READY_RELATIVE_PATH
    )


def test_semantic_rows_match_across_required_artifacts(tmp_path: Path) -> None:
    root = _isolated_root(tmp_path)
    record = publish_target_pipeline(root, now=FIXED_NOW)
    standardized = _read_csv(root / STANDARDIZED_RELATIVE_PATH)
    model_ready = _read_csv(root / MODEL_READY_RELATIVE_PATH)
    assert standardized == model_ready
    assert len(standardized) == 799
    assert record["checks"]["semantic_rows_match"] is True
    assert (root / FINDING_RELATIVE_PATH).is_file()
    assert (root / EVIDENCE_RELATIVE_PATH).is_file()


def test_raw_hash_unchanged_on_success(tmp_path: Path) -> None:
    root = _isolated_root(tmp_path)
    raw_path = root / RAW_RELATIVE_PATH
    before = sha256_file(raw_path)
    publish_target_pipeline(root, now=FIXED_NOW)
    assert sha256_file(raw_path) == before == EXPECTED_RAW_SHA256


def test_evidence_json_written_last_and_write_failure_cannot_pass(tmp_path: Path) -> None:
    root = _isolated_root(tmp_path)
    raw_path = root / RAW_RELATIVE_PATH
    before = sha256_file(raw_path)
    destinations: list[Path] = []
    staged_counts: list[int] = []

    def fail_final_evidence_write(staged: Path, destination: Path) -> None:
        staged_counts.append(len(list(root.rglob("*.tmp"))))
        destinations.append(destination)
        if destination == root / EVIDENCE_RELATIVE_PATH:
            raise OSError("forced final evidence failure")
        staged.replace(destination)

    with pytest.raises(PublicationError, match="forced final evidence failure"):
        publish_target_pipeline(
            root,
            replace_file=fail_final_evidence_write,
            now=FIXED_NOW,
        )

    assert destinations == [
        root / STANDARDIZED_RELATIVE_PATH,
        root / MODEL_READY_RELATIVE_PATH,
        root / FINDING_RELATIVE_PATH,
        root / EVIDENCE_RELATIVE_PATH,
    ]
    assert staged_counts == [4, 3, 2, 1]
    assert not (root / EVIDENCE_RELATIVE_PATH).exists()
    assert sha256_file(raw_path) == before == EXPECTED_RAW_SHA256
    assert not list(root.rglob("*.tmp"))


def test_roadmap_gate_rejects_missing_nonpass_and_stale_evidence(tmp_path: Path) -> None:
    missing_root = _isolated_root(tmp_path, "missing")
    with pytest.raises(RoadmapEligibilityError, match="unavailable"):
        assert_roadmap_completion_eligible(missing_root)

    nonpass_root = _isolated_root(tmp_path, "nonpass")
    publish_target_pipeline(nonpass_root, now=FIXED_NOW)
    evidence_path = nonpass_root / EVIDENCE_RELATIVE_PATH
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    evidence["classification"] = "fail"
    evidence_path.write_text(json.dumps(evidence), encoding="utf-8")
    with pytest.raises(RoadmapEligibilityError, match="not pass"):
        assert_roadmap_completion_eligible(nonpass_root)

    stale_root = _isolated_root(tmp_path, "stale")
    publish_target_pipeline(stale_root, now=FIXED_NOW)
    (stale_root / STANDARDIZED_RELATIVE_PATH).write_text("stale\n", encoding="utf-8")
    with pytest.raises(RoadmapEligibilityError, match="stale P1-02 evidence"):
        assert_roadmap_completion_eligible(stale_root)

    stale_model_root = _isolated_root(tmp_path, "stale-model")
    publish_target_pipeline(stale_model_root, now=FIXED_NOW)
    (stale_model_root / MODEL_READY_RELATIVE_PATH).write_text("stale\n", encoding="utf-8")
    with pytest.raises(RoadmapEligibilityError, match="stale P1-02 evidence"):
        assert_roadmap_completion_eligible(stale_model_root)


def test_roadmap_gate_rejects_rehashed_malformed_artifact(tmp_path: Path) -> None:
    root = _isolated_root(tmp_path)
    publish_target_pipeline(root, now=FIXED_NOW)
    standardized_path = root / STANDARDIZED_RELATIVE_PATH
    standardized_path.write_text(
        "unique_id,ds,y\nworld_bank_pink_sheet_monthly_arabica,2026-01-01,1\n",
        encoding="utf-8",
    )
    evidence_path = root / EVIDENCE_RELATIVE_PATH
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    evidence["standardized_sha256"] = sha256_file(standardized_path)
    evidence_path.write_text(json.dumps(evidence), encoding="utf-8")
    (root / FINDING_RELATIVE_PATH).write_text(
        target_pipeline.render_finding(evidence),
        encoding="utf-8",
    )
    with pytest.raises(RoadmapEligibilityError, match="monthly target contract"):
        assert_roadmap_completion_eligible(root)


def test_roadmap_gate_rejects_rehashed_semantic_drift(tmp_path: Path) -> None:
    root = _isolated_root(tmp_path)
    publish_target_pipeline(root, now=FIXED_NOW)
    model_ready_path = root / MODEL_READY_RELATIVE_PATH
    rows = _read_csv(model_ready_path)
    rows[-1]["y"] = "8.01"
    with model_ready_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(CSV_FIELDS), lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)
    evidence_path = root / EVIDENCE_RELATIVE_PATH
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    evidence["model_ready_sha256"] = sha256_file(model_ready_path)
    evidence_path.write_text(json.dumps(evidence), encoding="utf-8")
    (root / FINDING_RELATIVE_PATH).write_text(
        target_pipeline.render_finding(evidence),
        encoding="utf-8",
    )
    with pytest.raises(RoadmapEligibilityError, match="target rows differ"):
        assert_roadmap_completion_eligible(root)


def test_roadmap_gate_rejects_rehashed_artifacts_that_drift_from_source(
    tmp_path: Path,
) -> None:
    root = _isolated_root(tmp_path)
    publish_target_pipeline(root, now=FIXED_NOW)
    for relative_path in (STANDARDIZED_RELATIVE_PATH, MODEL_READY_RELATIVE_PATH):
        artifact_path = root / relative_path
        rows = _read_csv(artifact_path)
        rows[-1]["y"] = "8.01"
        with artifact_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=list(CSV_FIELDS), lineterminator="\n")
            writer.writeheader()
            writer.writerows(rows)

    evidence_path = root / EVIDENCE_RELATIVE_PATH
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    evidence["standardized_sha256"] = sha256_file(root / STANDARDIZED_RELATIVE_PATH)
    evidence["model_ready_sha256"] = sha256_file(root / MODEL_READY_RELATIVE_PATH)
    evidence_path.write_text(json.dumps(evidence), encoding="utf-8")
    (root / FINDING_RELATIVE_PATH).write_text(
        target_pipeline.render_finding(evidence),
        encoding="utf-8",
    )

    with pytest.raises(RoadmapEligibilityError, match="do not match the source workbook"):
        assert_roadmap_completion_eligible(root)


@pytest.mark.parametrize("classification", ["fail", "blocked", "unsupported"])
def test_nonpass_evidence_requires_unchecked_roadmap(
    tmp_path: Path,
    classification: str,
) -> None:
    root = _isolated_root(tmp_path)
    publish_target_pipeline(root, now=FIXED_NOW)
    evidence_path = root / EVIDENCE_RELATIVE_PATH
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    evidence["classification"] = classification
    evidence_path.write_text(json.dumps(evidence), encoding="utf-8")

    with pytest.raises(RoadmapEligibilityError, match="roadmap state"):
        assert_roadmap_consistent(root)

    roadmap_path = root / "docs" / "roadmap.md"
    roadmap_path.write_text(
        roadmap_path.read_text(encoding="utf-8").replace(
            "- [x] **P1-02",
            "- [ ] **P1-02",
        ),
        encoding="utf-8",
    )
    assert_roadmap_consistent(root)


def test_roadmap_gate_rejects_false_json_last_and_finding_drift(tmp_path: Path) -> None:
    false_flag_root = _isolated_root(tmp_path, "false-flag")
    publish_target_pipeline(false_flag_root, now=FIXED_NOW)
    evidence_path = false_flag_root / EVIDENCE_RELATIVE_PATH
    evidence = json.loads(evidence_path.read_text(encoding="utf-8"))
    evidence["checks"]["evidence_json_written_last"] = False
    evidence_path.write_text(json.dumps(evidence), encoding="utf-8")
    with pytest.raises(RoadmapEligibilityError, match="canonical P1-02 evidence is invalid"):
        assert_roadmap_completion_eligible(false_flag_root)

    finding_root = _isolated_root(tmp_path, "finding-drift")
    publish_target_pipeline(finding_root, now=FIXED_NOW)
    finding_path = finding_root / FINDING_RELATIVE_PATH
    finding_path.write_text(
        finding_path.read_text(encoding="utf-8") + "Unexpected text\n",
        encoding="utf-8",
    )
    with pytest.raises(RoadmapEligibilityError, match="finding does not match"):
        assert_roadmap_completion_eligible(finding_root)


def test_published_state_is_schema_valid_ordered_and_current(tmp_path: Path) -> None:
    root = _isolated_root(tmp_path)
    record = publish_target_pipeline(root, now=FIXED_NOW)
    validate_published_state(root, record)
    rows = _read_csv(root / STANDARDIZED_RELATIVE_PATH)
    dates = [date.fromisoformat(row["ds"]) for row in rows]
    assert dates == sorted(set(dates))
    assert dates[0] == date(1960, 1, 1)
    assert dates[-1] == date(2026, 7, 1)
    assert len(dates) == 799


def test_cli_requires_absolute_repo_root() -> None:
    assert target_pipeline.main(["publish", "--repo-root", "relative/repo"]) == 1
