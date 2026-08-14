"""Build the monthly Arabica target artifacts from the preserved workbook."""

from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import os
import re
import sys
import tempfile
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable, Mapping, Sequence

from .contracts import (
    EXPECTED_PERIOD_COUNT,
    EXPECTED_PERIOD_END,
    EXPECTED_PERIOD_START,
    EXPECTED_RAW_SHA256,
    PUBLICATION_POLICY,
    TARGET_COLUMN,
    WORKSHEET_NAME,
)
from .evidence import EvidenceError, assert_finding_matches_record, validate_evidence_record
from .paths import (
    dependency_readiness_evidence_path,
    dependency_readiness_finding_path,
    raw_workbook_path,
)
from .readiness import (
    ReadinessError,
    assert_latest_run_roadmap_consistency,
    parse_month_token,
    sha256_file,
)

TASK_ID = "P1-02"
EVIDENCE_SCHEMA_VERSION = 1
UNIQUE_ID = "world_bank_pink_sheet_monthly_arabica"
CSV_FIELDS = ("unique_id", "ds", "y")
RAW_RELATIVE_PATH = Path(
    "data/raw/world_bank/pink_sheet/CMO-Historical-Data-Monthly.xlsx"
)
STANDARDIZED_RELATIVE_PATH = Path(
    "data/standardized/world_bank/pink_sheet/coffee_arabica_monthly.csv"
)
MODEL_READY_RELATIVE_PATH = Path(
    "data/model_ready/world_bank_pink_sheet_monthly_arabica/target.csv"
)
FINDING_RELATIVE_PATH = Path("docs/findings/phase1/target_pipeline.md")
EVIDENCE_RELATIVE_PATH = Path("docs/findings/phase1/evidence/target_pipeline.json")
P1_02_ROADMAP_PATTERN = re.compile(r"(?m)^- \[([ x-])\] \*\*P1-02\b.*$")
EXPECTED_HEADER_ROW = 5
EXPECTED_TARGET_COLUMN_INDEX = 13
MONTH_LIKE_TOKEN = re.compile(r"^\d{4}M")
SHA256_PATTERN = re.compile(r"^[0-9a-f]{64}$")

REQUIRED_CHECKS = (
    "p101_evidence_valid",
    "p101_evidence_pass",
    "p101_finding_matches",
    "p101_publication_policy_matches",
    "source_exact_target_found",
    "row_count_799",
    "monthly_contiguous",
    "numeric_serializer_valid",
    "semantic_rows_match",
    "schemas_valid",
    "raw_hash_unchanged",
    "publication_bytes_valid",
    "evidence_json_written_last",
)

REQUIRED_EVIDENCE_FIELDS = {
    "schema_version",
    "task_id",
    "run_id",
    "timestamp_utc",
    "classification",
    "source_workbook_path",
    "source_sha256_before",
    "source_sha256_after",
    "sheet_name",
    "target_column",
    "header_row",
    "target_column_index",
    "period_start",
    "period_end",
    "period_count",
    "unique_id",
    "standardized_path",
    "model_ready_path",
    "standardized_sha256",
    "model_ready_sha256",
    "schemas",
    "publication_policy",
    "checks",
    "errors",
}


class TargetPipelineError(RuntimeError):
    """Base error for target-pipeline contract failures."""


class DependencyGateError(TargetPipelineError):
    """Raised when current P1-01 evidence is not a canonical pass."""


class TargetSelectionError(TargetPipelineError):
    """Raised when the exact target column cannot be selected once."""


class InvalidTimestampError(TargetPipelineError):
    """Raised when a monthly timestamp violates the source contract."""


class InvalidTargetValueError(TargetPipelineError):
    """Raised when a target cell is not a finite workbook numeric."""


class TargetEvidenceError(TargetPipelineError):
    """Raised when P1-02 evidence violates its compact schema."""


class PublicationError(TargetPipelineError):
    """Raised when staged artifacts cannot be published."""


class RoadmapEligibilityError(TargetPipelineError):
    """Raised when P1-02 evidence cannot authorize roadmap completion."""


@dataclass(frozen=True)
class TargetRow:
    unique_id: str
    ds: date
    y: str


@dataclass(frozen=True)
class TargetObservation:
    rows: tuple[TargetRow, ...]
    header_row: int
    target_column_index: int
    period_start: str
    period_end: str
    source_sha256_before: str
    source_sha256_after: str


ReplaceFile = Callable[[Path, Path], None]


def serialize_numeric(value: object) -> str:
    """Serialize a finite workbook numeric without imposing display formatting."""

    if type(value) not in {int, float}:
        raise InvalidTargetValueError("target value must be a finite int or float")
    if isinstance(value, float) and not math.isfinite(value):
        raise InvalidTargetValueError("target value must be finite")
    try:
        decimal_value = Decimal(str(value))
    except InvalidOperation as exc:
        raise InvalidTargetValueError("target value is not a valid decimal") from exc
    if not decimal_value.is_finite():
        raise InvalidTargetValueError("target value must be finite")
    return format(decimal_value, "f")


def _validate_serialized_numeric(value: str) -> None:
    if not value:
        raise InvalidTargetValueError("serialized target must be non-empty")
    try:
        decimal_value = Decimal(value)
    except InvalidOperation as exc:
        raise InvalidTargetValueError("serialized target is not numeric") from exc
    if not decimal_value.is_finite() or format(decimal_value, "f") != value:
        raise InvalidTargetValueError("serialized target must be a canonical finite decimal")


def _next_month(value: date) -> date:
    if value.month == 12:
        return date(value.year + 1, 1, 1)
    return date(value.year, value.month + 1, 1)


def validate_rows(rows: Sequence[TargetRow]) -> None:
    if len(rows) != EXPECTED_PERIOD_COUNT:
        raise InvalidTimestampError(
            f"expected {EXPECTED_PERIOD_COUNT} monthly rows, found {len(rows)}"
        )
    dates = [row.ds for row in rows]
    if len(set(dates)) != len(dates) or dates != sorted(dates):
        raise InvalidTimestampError("monthly timestamps must be unique and strictly ordered")
    for previous, current in zip(dates, dates[1:]):
        if current != _next_month(previous):
            raise InvalidTimestampError("monthly timestamps must be contiguous calendar months")
    if f"{dates[0].year:04d}M{dates[0].month:02d}" != EXPECTED_PERIOD_START:
        raise InvalidTimestampError(f"first period must be {EXPECTED_PERIOD_START}")
    if f"{dates[-1].year:04d}M{dates[-1].month:02d}" != EXPECTED_PERIOD_END:
        raise InvalidTimestampError(f"last period must be {EXPECTED_PERIOD_END}")
    if any(row.unique_id != UNIQUE_ID for row in rows):
        raise TargetPipelineError("every row must use the canonical unique_id")
    for row in rows:
        _validate_serialized_numeric(row.y)


def _find_target_header(worksheet: Any) -> tuple[int, int]:
    positions: list[tuple[int, int]] = []
    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        for column_index, value in enumerate(row, start=1):
            if value == TARGET_COLUMN:
                positions.append((row_index, column_index))
    if not positions:
        raise TargetSelectionError(f"target column not found: {TARGET_COLUMN}")
    if len(positions) != 1:
        raise TargetSelectionError(
            f"target column must appear exactly once, found {len(positions)}"
        )
    return positions[0]


def _read_target_rows(
    worksheet: Any,
    *,
    header_row: int,
    target_column_index: int,
) -> tuple[TargetRow, ...]:
    rows: list[TargetRow] = []
    periods_started = False
    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        period_value = row[0] if row else None
        if isinstance(period_value, str) and MONTH_LIKE_TOKEN.match(period_value):
            try:
                period = parse_month_token(period_value)
            except ReadinessError as exc:
                raise InvalidTimestampError(f"invalid monthly timestamp: {period_value!r}") from exc
            periods_started = True
            target_value = row[target_column_index - 1]
            rows.append(TargetRow(UNIQUE_ID, period, serialize_numeric(target_value)))
        elif periods_started and period_value not in {None, ""}:
            raise InvalidTimestampError(
                f"invalid monthly timestamp after data began: {period_value!r}"
            )
    if not rows:
        raise InvalidTimestampError("workbook contains no monthly timestamp rows")
    return tuple(rows)


def extract_target(workbook_path: Path) -> TargetObservation:
    """Read the exact monthly Arabica target without mutating the workbook."""

    if not workbook_path.is_file():
        raise TargetPipelineError(f"raw workbook does not exist: {workbook_path}")
    source_sha256_before = sha256_file(workbook_path)
    if source_sha256_before != EXPECTED_RAW_SHA256:
        raise TargetPipelineError("raw workbook hash does not match the preserved source")

    workbook: Any | None = None
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        if WORKSHEET_NAME not in workbook.sheetnames:
            raise TargetSelectionError(f"worksheet not found: {WORKSHEET_NAME}")
        worksheet = workbook[WORKSHEET_NAME]
        header_row, target_column_index = _find_target_header(worksheet)
        rows = _read_target_rows(
            worksheet,
            header_row=header_row,
            target_column_index=target_column_index,
        )
    finally:
        if workbook is not None:
            workbook.close()

    source_sha256_after = sha256_file(workbook_path)
    if source_sha256_after != source_sha256_before:
        raise TargetPipelineError("raw workbook hash changed during target extraction")
    validate_rows(rows)
    first, last = rows[0].ds, rows[-1].ds
    return TargetObservation(
        rows=rows,
        header_row=header_row,
        target_column_index=target_column_index,
        period_start=f"{first.year:04d}M{first.month:02d}",
        period_end=f"{last.year:04d}M{last.month:02d}",
        source_sha256_before=source_sha256_before,
        source_sha256_after=source_sha256_after,
    )


def serialize_csv(rows: Sequence[TargetRow]) -> bytes:
    """Return deterministic target-only CSV bytes."""

    stream = io.StringIO(newline="")
    writer = csv.writer(stream, lineterminator="\n")
    writer.writerow(CSV_FIELDS)
    for row in rows:
        writer.writerow((row.unique_id, row.ds.isoformat(), row.y))
    return stream.getvalue().encode("utf-8")


def parse_target_csv(path: Path) -> tuple[TargetRow, ...]:
    """Parse and validate a published target CSV independently of its evidence hash."""

    try:
        content = path.read_bytes()
        text = content.decode("utf-8")
    except (OSError, UnicodeDecodeError) as exc:
        raise RoadmapEligibilityError(f"target CSV is unreadable: {path}") from exc
    reader = csv.reader(io.StringIO(text, newline=""))
    try:
        header = next(reader)
    except StopIteration as exc:
        raise RoadmapEligibilityError(f"target CSV is empty: {path}") from exc
    if header != list(CSV_FIELDS):
        raise RoadmapEligibilityError(f"target CSV has the wrong schema: {path}")
    rows: list[TargetRow] = []
    try:
        for values in reader:
            if len(values) != len(CSV_FIELDS):
                raise RoadmapEligibilityError(f"target CSV row has the wrong width: {path}")
            unique_id, ds_text, y = values
            ds = date.fromisoformat(ds_text)
            if ds.isoformat() != ds_text or ds.day != 1:
                raise InvalidTimestampError("published ds must be a first-of-month ISO date")
            _validate_serialized_numeric(y)
            rows.append(TargetRow(unique_id, ds, y))
        validate_rows(rows)
    except (ValueError, TargetPipelineError) as exc:
        raise RoadmapEligibilityError(f"target CSV violates the monthly target contract: {path}") from exc
    parsed = tuple(rows)
    if serialize_csv(parsed) != content:
        raise RoadmapEligibilityError(f"target CSV serialization is not canonical: {path}")
    return parsed


def sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _utc_now() -> datetime:
    return datetime.now(timezone.utc).replace(microsecond=0)


def _timestamp_text(value: datetime) -> str:
    return value.isoformat().replace("+00:00", "Z")


def _run_id(value: datetime) -> str:
    return f"P1-02-{value.strftime('%Y%m%dT%H%M%SZ')}"


def validate_target_evidence(record: dict[str, Any]) -> None:
    """Validate the compact P1-02 evidence schema and pass invariants."""

    fields = set(record)
    allowed_fields = REQUIRED_EVIDENCE_FIELDS | {"notes"}
    missing = sorted(REQUIRED_EVIDENCE_FIELDS - fields)
    extra = sorted(fields - allowed_fields)
    if missing:
        raise TargetEvidenceError(f"missing evidence fields: {', '.join(missing)}")
    if extra:
        raise TargetEvidenceError(f"unsupported evidence fields: {', '.join(extra)}")
    if record["schema_version"] != EVIDENCE_SCHEMA_VERSION or record["task_id"] != TASK_ID:
        raise TargetEvidenceError("evidence identity does not match P1-02 schema version 1")
    if not isinstance(record["run_id"], str) or not record["run_id"].strip():
        raise TargetEvidenceError("run_id must be non-empty")
    timestamp = record["timestamp_utc"]
    if not isinstance(timestamp, str) or not timestamp.endswith("Z"):
        raise TargetEvidenceError("timestamp_utc must be an ISO-8601 UTC value ending in Z")
    try:
        datetime.fromisoformat(timestamp.replace("Z", "+00:00"))
    except ValueError as exc:
        raise TargetEvidenceError("timestamp_utc is malformed") from exc
    if record["classification"] not in {"pass", "fail", "blocked", "unsupported"}:
        raise TargetEvidenceError("unsupported evidence classification")
    if record["publication_policy"] != PUBLICATION_POLICY.as_dict():
        raise TargetEvidenceError("publication policy does not match P1-01")
    expected_schemas = {
        "standardized": list(CSV_FIELDS),
        "model_ready": list(CSV_FIELDS),
    }
    if record["schemas"] != expected_schemas:
        raise TargetEvidenceError("artifact schemas must be target-only and ordered")
    checks = record["checks"]
    if not isinstance(checks, dict) or set(checks) != set(REQUIRED_CHECKS):
        raise TargetEvidenceError("checks must contain exactly the required named checks")
    if any(type(value) is not bool for value in checks.values()):
        raise TargetEvidenceError("all named checks must be boolean")
    if not isinstance(record["errors"], list) or any(
        not isinstance(error, str) for error in record["errors"]
    ):
        raise TargetEvidenceError("errors must be a list of strings")
    for field in ("source_sha256_before", "source_sha256_after", "standardized_sha256", "model_ready_sha256"):
        if not isinstance(record[field], str) or SHA256_PATTERN.fullmatch(record[field]) is None:
            raise TargetEvidenceError(f"{field} must be a lowercase SHA-256")

    if record["classification"] == "pass":
        if not all(checks.values()) or record["errors"]:
            raise TargetEvidenceError("pass evidence requires true checks and no errors")
        if not (
            record["source_sha256_before"]
            == record["source_sha256_after"]
            == EXPECTED_RAW_SHA256
        ):
            raise TargetEvidenceError("pass evidence must preserve the expected raw hash")
        if record["sheet_name"] != WORKSHEET_NAME or record["target_column"] != TARGET_COLUMN:
            raise TargetEvidenceError("pass evidence must record the exact source selection")
        if record["source_workbook_path"] != RAW_RELATIVE_PATH.as_posix():
            raise TargetEvidenceError("pass evidence must record the canonical source workbook")
        if (
            record["header_row"] != EXPECTED_HEADER_ROW
            or record["target_column_index"] != EXPECTED_TARGET_COLUMN_INDEX
        ):
            raise TargetEvidenceError("pass evidence must record the canonical header and target index")
        if (
            record["period_start"] != EXPECTED_PERIOD_START
            or record["period_end"] != EXPECTED_PERIOD_END
            or record["period_count"] != EXPECTED_PERIOD_COUNT
        ):
            raise TargetEvidenceError("pass evidence must record the expected monthly extent")
        if record["unique_id"] != UNIQUE_ID:
            raise TargetEvidenceError("pass evidence must record the canonical unique_id")
        if record["standardized_path"] != STANDARDIZED_RELATIVE_PATH.as_posix():
            raise TargetEvidenceError("pass evidence has the wrong standardized path")
        if record["model_ready_path"] != MODEL_READY_RELATIVE_PATH.as_posix():
            raise TargetEvidenceError("pass evidence has the wrong model-ready path")


def _load_p101_gate(repo_root: Path) -> dict[str, Any]:
    evidence_path = dependency_readiness_evidence_path(repo_root)
    try:
        raw = json.loads(evidence_path.read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            raise EvidenceError("canonical P1-01 evidence must be an object")
        validate_evidence_record(raw)
        if raw["classification"] != "pass":
            raise EvidenceError("canonical P1-01 evidence is not pass")
        assert_finding_matches_record(dependency_readiness_finding_path(repo_root), raw)
        assert_latest_run_roadmap_consistency(repo_root)
    except (OSError, json.JSONDecodeError, EvidenceError, ReadinessError) as exc:
        raise DependencyGateError(f"P1-01 gate is not a current canonical pass: {exc}") from exc
    return raw


def _render_finding_fields(values: Mapping[str, Any]) -> str:
    return (
        "# Phase 1 Monthly Target Pipeline\n\n"
        f"- Run ID: `{values['run_id']}`\n"
        f"- Classification: `{values['classification']}`\n"
        f"- Source: `{values['sheet_name']}` / `{values['target_column']}`\n"
        f"- Raw SHA-256 before: `{values['source_sha256_before']}`\n"
        f"- Raw SHA-256 after: `{values['source_sha256_after']}`\n"
        f"- Monthly extent: `{values['period_start']}` through `{values['period_end']}` "
        f"(`{values['period_count']}` rows)\n"
        f"- Schema: `{','.join(values['schemas']['standardized'])}`\n"
        f"- Standardized SHA-256: `{values['standardized_sha256']}`\n"
        f"- Model-ready SHA-256: `{values['model_ready_sha256']}`\n\n"
        "## Publication-availability policy\n\n"
        f"- Evaluation label: `{PUBLICATION_POLICY.evaluation_label}`\n"
        f"- Availability proxy: `{PUBLICATION_POLICY.availability_proxy}`\n"
        f"- Limitation: {PUBLICATION_POLICY.limitation}\n"
        f"- Prohibited claim: {PUBLICATION_POLICY.prohibited_claim}\n\n"
        "## Transform boundary\n\n"
        "The artifact contains only the observed monthly target level. No fill, interpolation, "
        "resampling, frequency conversion, scaling, differencing, covariates, engineered features, "
        "or future values were used.\n\n"
        "Machine-readable evidence: `docs/findings/phase1/evidence/target_pipeline.json`\n"
    )


def render_finding(record: dict[str, Any]) -> str:
    validate_target_evidence(record)
    return _render_finding_fields(record)


def _build_record(
    observation: TargetObservation,
    *,
    standardized_sha256: str,
    model_ready_sha256: str,
    now: datetime,
) -> dict[str, Any]:
    return {
        "schema_version": EVIDENCE_SCHEMA_VERSION,
        "task_id": TASK_ID,
        "run_id": _run_id(now),
        "timestamp_utc": _timestamp_text(now),
        "classification": "pass",
        "source_workbook_path": RAW_RELATIVE_PATH.as_posix(),
        "source_sha256_before": observation.source_sha256_before,
        "source_sha256_after": observation.source_sha256_after,
        "sheet_name": WORKSHEET_NAME,
        "target_column": TARGET_COLUMN,
        "header_row": observation.header_row,
        "target_column_index": observation.target_column_index,
        "period_start": observation.period_start,
        "period_end": observation.period_end,
        "period_count": len(observation.rows),
        "unique_id": UNIQUE_ID,
        "standardized_path": STANDARDIZED_RELATIVE_PATH.as_posix(),
        "model_ready_path": MODEL_READY_RELATIVE_PATH.as_posix(),
        "standardized_sha256": standardized_sha256,
        "model_ready_sha256": model_ready_sha256,
        "schemas": {
            "standardized": list(CSV_FIELDS),
            "model_ready": list(CSV_FIELDS),
        },
        "publication_policy": PUBLICATION_POLICY.as_dict(),
        "checks": {name: True for name in REQUIRED_CHECKS},
        "errors": [],
    }


def _stage_bytes(destination: Path, content: bytes) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    temporary_path: Path | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="wb",
            dir=destination.parent,
            prefix=f".{destination.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            handle.write(content)
            handle.flush()
            os.fsync(handle.fileno())
            temporary_path = Path(handle.name)
    except Exception:
        if temporary_path is not None and temporary_path.exists():
            temporary_path.unlink()
        raise
    if temporary_path is None:
        raise PublicationError(f"could not stage artifact: {destination}")
    return temporary_path


def _replace_file(staged: Path, destination: Path) -> None:
    staged.replace(destination)


def _publish_buffers(
    buffers: Sequence[tuple[Path, bytes]],
    *,
    replace_file: ReplaceFile,
) -> None:
    staged: list[tuple[Path, Path]] = []
    try:
        for destination, content in buffers:
            staged.append((_stage_bytes(destination, content), destination))
        for temporary_path, destination in staged:
            replace_file(temporary_path, destination)
    except OSError as exc:
        raise PublicationError(f"artifact publication failed: {exc}") from exc
    finally:
        for temporary_path, _ in staged:
            if temporary_path.exists():
                temporary_path.unlink()


def publish_target_pipeline(
    repo_root: Path,
    *,
    replace_file: ReplaceFile = _replace_file,
    now: datetime | None = None,
) -> dict[str, Any]:
    """Validate P1-01, build all target bytes, and publish JSON last."""

    root = repo_root.resolve()
    _load_p101_gate(root)
    workbook_path = raw_workbook_path(root)
    raw_before = sha256_file(workbook_path)
    try:
        observation = extract_target(workbook_path)
        csv_bytes = serialize_csv(observation.rows)
        standardized_sha256 = sha256_bytes(csv_bytes)
        model_ready_sha256 = sha256_bytes(csv_bytes)
        publication_time = now or _utc_now()
        finding_fields = {
            "run_id": _run_id(publication_time),
            "classification": "pass",
            "sheet_name": WORKSHEET_NAME,
            "target_column": TARGET_COLUMN,
            "source_sha256_before": observation.source_sha256_before,
            "source_sha256_after": observation.source_sha256_after,
            "period_start": observation.period_start,
            "period_end": observation.period_end,
            "period_count": len(observation.rows),
            "schemas": {
                "standardized": list(CSV_FIELDS),
                "model_ready": list(CSV_FIELDS),
            },
            "standardized_sha256": standardized_sha256,
            "model_ready_sha256": model_ready_sha256,
        }
        finding_bytes = _render_finding_fields(finding_fields).encode("utf-8")
        record = _build_record(
            observation,
            standardized_sha256=standardized_sha256,
            model_ready_sha256=model_ready_sha256,
            now=publication_time,
        )
        validate_target_evidence(record)
        if render_finding(record).encode("utf-8") != finding_bytes:
            raise TargetEvidenceError("finding buffer does not match final evidence")
        evidence_bytes = (json.dumps(record, indent=2, sort_keys=True) + "\n").encode("utf-8")
        if json.loads(evidence_bytes) != record:
            raise TargetEvidenceError("serialized evidence does not round-trip")
        buffers = (
            (root / STANDARDIZED_RELATIVE_PATH, csv_bytes),
            (root / MODEL_READY_RELATIVE_PATH, csv_bytes),
            (root / FINDING_RELATIVE_PATH, finding_bytes),
            (root / EVIDENCE_RELATIVE_PATH, evidence_bytes),
        )
        _publish_buffers(buffers, replace_file=replace_file)
    except Exception as exc:
        if workbook_path.is_file() and sha256_file(workbook_path) != raw_before:
            raise TargetPipelineError("raw workbook hash changed on a failed pipeline run") from exc
        raise
    if sha256_file(workbook_path) != raw_before:
        raise TargetPipelineError("raw workbook hash changed during publication")
    validate_published_state(root)
    return record


def _load_target_evidence(repo_root: Path) -> dict[str, Any]:
    path = repo_root / EVIDENCE_RELATIVE_PATH
    try:
        raw = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise RoadmapEligibilityError(f"canonical P1-02 evidence is unavailable: {exc}") from exc
    if not isinstance(raw, dict):
        raise RoadmapEligibilityError("canonical P1-02 evidence must be an object")
    try:
        validate_target_evidence(raw)
    except TargetEvidenceError as exc:
        raise RoadmapEligibilityError(f"canonical P1-02 evidence is invalid: {exc}") from exc
    return raw


def validate_published_state(repo_root: Path, record: dict[str, Any] | None = None) -> None:
    """Validate canonical evidence against the currently published CSV artifacts."""

    root = repo_root.resolve()
    current = record if record is not None else _load_target_evidence(root)
    validate_target_evidence(current)
    if current["classification"] != "pass":
        raise RoadmapEligibilityError("P1-02 evidence is not pass")
    source_path = root / current["source_workbook_path"]
    if not source_path.is_file() or sha256_file(source_path) != EXPECTED_RAW_SHA256:
        raise RoadmapEligibilityError("canonical raw workbook provenance does not match evidence")

    artifact_rows: list[tuple[TargetRow, ...]] = []
    for path_field, hash_field in (
        ("standardized_path", "standardized_sha256"),
        ("model_ready_path", "model_ready_sha256"),
    ):
        path = root / current[path_field]
        if not path.is_file() or sha256_file(path) != current[hash_field]:
            raise RoadmapEligibilityError(f"stale P1-02 evidence for {path_field}")
        artifact_rows.append(parse_target_csv(path))
    if artifact_rows[0] != artifact_rows[1]:
        raise RoadmapEligibilityError("standardized and model-ready target rows differ")
    try:
        source_rows = extract_target(source_path).rows
    except TargetPipelineError as exc:
        raise RoadmapEligibilityError(f"canonical source target is invalid: {exc}") from exc
    if artifact_rows[0] != source_rows:
        raise RoadmapEligibilityError("published target rows do not match the source workbook")
    finding = root / FINDING_RELATIVE_PATH
    if not finding.is_file():
        raise RoadmapEligibilityError("P1-02 finding does not exist")
    if finding.read_text(encoding="utf-8") != render_finding(current):
        raise RoadmapEligibilityError("P1-02 finding does not match canonical evidence")


def assert_roadmap_completion_eligible(repo_root: Path) -> None:
    """Require current passing artifact hashes before the P1-02 checkbox can be checked."""

    root = repo_root.resolve()
    _load_p101_gate(root)
    validate_published_state(root)


def assert_roadmap_consistent(repo_root: Path) -> None:
    root = repo_root.resolve()
    evidence = _load_target_evidence(root)
    roadmap = (root / "docs" / "roadmap.md").read_text(encoding="utf-8")
    matches = P1_02_ROADMAP_PATTERN.findall(roadmap)
    expected_state = "x" if evidence["classification"] == "pass" else " "
    if matches != [expected_state]:
        raise RoadmapEligibilityError("P1-02 roadmap state does not match canonical evidence")
    if evidence["classification"] == "pass":
        validate_published_state(root, evidence)


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)
    publish = subparsers.add_parser("publish", help="publish the deterministic P1-02 artifacts")
    publish.add_argument("--repo-root", type=Path, required=True)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    arguments = _parser().parse_args(argv)
    if not arguments.repo_root.is_absolute():
        print("--repo-root must be an absolute path", file=sys.stderr)
        return 1
    try:
        record = publish_target_pipeline(arguments.repo_root)
    except TargetPipelineError as exc:
        print(str(exc), file=sys.stderr)
        return 1
    print(json.dumps(record, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
